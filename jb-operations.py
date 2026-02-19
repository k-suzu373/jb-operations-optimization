from __future__ import annotations
import argparse
import datetime as dt
import re
import random
from dataclasses import dataclass
from typing import Optional, Dict, Any, Tuple, List
import pandas as pd
import openpyxl
from ortools.sat.python import cp_model
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

LOC_CODE_MAP = {
    # 文字列でも変換可能にするための対応表
    "三鷹": "JB01",
    "Mitaka": "JB01",
    "中野": "JB07",
    "Nakano": "JB07",
    "御茶ノ水": "JB18",
    "Ochanomizu": "JB18",
    "千葉": "JB39",
    "Chiba": "JB39",
}


# -------I/O helpers----------
def _require_columns(df: pd.DataFrame, required: List[str], sheet: str) -> None:
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"[{sheet}] 必須列が見つかりません: {missing}\n"
            f"現在の列: {list(df.columns)}"
        )


def to_station_code(x):
    # 文字列駅をナンバリングに変換
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return s
    if re.fullmatch(r"JB\d{2}", s):
        return s
    return LOC_CODE_MAP.get(s, s)


# 塗りつぶし色
FILL_MITAKA = PatternFill("solid", fgColor="4F82EA")  # 紺
FILL_NAKANO = PatternFill("solid", fgColor="9DC3E6")  # 水色
FILL_OCHA = PatternFill("solid", fgColor="F4B084")  # オレンジ
FILL_CHIBA = PatternFill("solid", fgColor="da70d6")  # 紫
FILL_OTHER = PatternFill("solid", fgColor="FFF2CC")  # 黄色
FONT_WHITE = Font(color="FFFFFF")


def station_fill(value):
    # セル値から背景色を決める（4駅=指定色、それ以外=黄色）
    if value is None:
        return None
    code = to_station_code(value)
    if code == "JB01":
        return FILL_MITAKA
    elif code == "JB07":
        return FILL_NAKANO
    elif code == "JB18":
        return FILL_OCHA
    elif code == "JB39":
        return FILL_CHIBA
    else:
        return FILL_OTHER


def set_white_font(cell, value):
    """三鷹(JB01)だけ文字を白にする"""
    if to_station_code(value) == "JB01":
        cell.font = FONT_WHITE


def _normalize_rule_key(key: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(key).lower())


def _parse_rules(rules: pd.DataFrame) -> Dict[str, int]:
    lower_cols = {str(c).lower(): c for c in rules.columns}
    if "key" in lower_cols and "value" in lower_cols:
        key_col = lower_cols["key"]
        value_col = lower_cols["value"]
        raw = {
            str(k).strip(): v
            for k, v in zip(rules[key_col].tolist(), rules[value_col].tolist())
        }
    else:
        raw = rules.iloc[0].to_dict()

    key_map = {
        "max_days_since_inspectiona": "max_days_since_inspectionA",
        "maxdayssinceinspectiona": "max_days_since_inspectionA",
        "max_days_since_inspectionb": "max_days_since_inspectionB",
        "maxdayssinceinspectionb": "max_days_since_inspectionB",
        "max_distance_since_inspectionb": "max_distance_since_inspectionB",
        "maxdistancesinceinspectionb": "max_distance_since_inspectionB",
        "max_distance_since_inspectionb_km": "max_distance_since_inspectionB",
        "maxdistancesinceinspectionbkm": "max_distance_since_inspectionB",
        "max_distance_since_inspectionbkm": "max_distance_since_inspectionB",
    }

    parsed: Dict[str, int] = {}
    for raw_key, value in raw.items():
        normalized = _normalize_rule_key(raw_key)
        if normalized in key_map:
            parsed[key_map[normalized]] = int(value)

    required = [
        "max_days_since_inspectionA",
        "max_days_since_inspectionB",
        "max_distance_since_inspectionB",
    ]
    missing = [k for k in required if k not in parsed]
    if missing:
        raise ValueError(
            "[rules] 必須キーが見つかりません: "
            f"{missing} (存在キー: {list(raw.keys())})"
        )
    return parsed


def read_master(master_path: str) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, int]]:
    ops = pd.read_excel(master_path, sheet_name="operations")
    forms = pd.read_excel(master_path, sheet_name="formations")
    rules = pd.read_excel(master_path, sheet_name="rules")

    # 想定する最低限の列
    _require_columns(
        ops,
        required=[
            "operation_id",
            "required",
            "is_inspection_B",
            "start_loc",
            "end_loc",
            "distance_km",
        ],
        sheet="operations",
    )  # 運用

    _require_columns(
        forms,
        required=[
            "formation_id",
            "init_location",
            "init_days_since_inspectionA",
            "init_days_since_inspectionB",
        ],
        sheet="formations",
    )  # 編成

    if rules.empty:
        raise ValueError("[rules] ルールシートが空です")

    # IDは文字列に寄せる（Excel側で数字扱いでも安定）
    # map関数でリストを文字列化したものを新しいリストに更新
    ops["start_loc"] = ops["start_loc"].map(to_station_code)
    ops["end_loc"] = ops["end_loc"].map(to_station_code)
    forms["init_location"] = forms["init_location"].map(to_station_code)

    if ops["distance_km"].isna().any():
        raise ValueError("[operations] distance_km に欠損があります")
    for opt_col in [
        "init_distance_since_inspectionB_km",
        "init_total_distance_km",
        "distance_km",
    ]:
        if opt_col in forms.columns and forms[opt_col].isna().any():
            raise ValueError(f"[formations] {opt_col} に欠損があります")

    rules_row = _parse_rules(rules)
    return (
        ops,
        forms,
        rules_row,
    )  # 検査までの日数設定


# ---------- Core logic (baseline allocator) ----------


@dataclass
class FormationState:
    loc: Any
    daysA: int
    daysB: int
    distB_km: int
    total_km: int
    prev_op_id: Optional[str] = None  # 前日に担当した運用ID


NON_DEADHEAD_LOC_JUMPS = {
    ("JB33", "JB30"),
    ("JB33", "JB35"),
}  # 回送扱いしない駅の組み合わせ

NON_DEADHEAD_OP_CHAINS = {
    ("53B", "55B"),
}  # 回送扱いしない運用の組み合わせ

FIXED_NEXT_OP = {
    "47B": "49B",
    "53B": "55B",
    "09B": "11B",
    "81B": "83B",
}  # 翌日固定になる運用の組み合わせ

# 目的関数の重み（range最優先、deadheadとoverdueは補助的に抑える）
W_RANGE = 1000
W_DEADHEAD = 10
W_OVD = 100


def make_baseline_schedule(
    ops: pd.DataFrame,
    forms: pd.DataFrame,
    max_days_since_inspectionA: int,
    max_days_since_inspectionB: int,
    max_distance_since_inspectionB: int,
    days: int = 30,
    start_date: Optional[str] = None,
    default_idle_op: str = "IDOL_Mitaka",
    seed: int = 30,
    tsudanuma_code: str = "JB33",
    priority_start_codes: Optional[List[str]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    制約を満たす割当：
      - required=1 の運用を毎日埋める（残りrequiredはCP-SATで割当）
      - B検査成立時はdistB_kmを0にリセット
      - distB_kmが上限を超えないように割当
      - 走行距離の偏り（range）とdeadhead、期限超過を目的関数で抑制
      - 余り編成は現在位置に対応する IDOL_* を割当（なければ default_idle_op）
    """

    ops = ops.copy()
    ops["operation_id"] = ops["operation_id"].astype(str)
    required_ops = ops[ops["required"] == 1].copy().reset_index(drop=True)
    required_ops["operation_id"] = required_ops["operation_id"].astype(str)

    if priority_start_codes is None:
        priority_start_codes = ["JB01", "JB07", "JB18", "JB39"]  # 津田沼以外の優先駅

    inspection_a_start_codes = {"JB01", "JB07", "JB33"}  # A検査可能駅

    # 予備待機（required=0 かつ 検査Bじゃないもの）から start_loc→operation_id を作る
    idle_ops = ops[(ops["required"] == 0) & (ops["is_inspection_B"] == 0)].copy()
    idle_by_start: Dict[Any, str] = dict(
        zip(idle_ops["start_loc"], idle_ops["operation_id"])
    )

    # operation_id→行の引き当て用（終点などを取る）
    op_by_id = {str(row.operation_id): row for row in ops.itertuples(index=False)}

    # 編成状態
    state: Dict[str, FormationState] = {}
    for r in forms.itertuples(index=False):
        formation_distance = getattr(r, "distance_km", None)
        if formation_distance is None:
            formation_distance = getattr(r, "init_total_distance_km", 0)
        state[str(r.formation_id)] = FormationState(
            loc=r.init_location,
            daysA=int(r.init_days_since_inspectionA),
            daysB=int(r.init_days_since_inspectionB),
            distB_km=int(getattr(r, "init_distance_since_inspectionB_km", 0) or 0),
            total_km=int(formation_distance or 0),
            prev_op_id=None,
        )

    for formation_id, f_state in state.items():
        if f_state.distB_km > max_distance_since_inspectionB:
            raise ValueError(
                f"[init] distB超過: formation={formation_id}, "
                f"distB_km={f_state.distB_km}, "
                f"limit={max_distance_since_inspectionB}"
            )

    formation_ids = list(forms["formation_id"].astype(str))

    # 日ラベル
    if start_date:
        start = dt.date.fromisoformat(start_date)
        day_labels = [start + dt.timedelta(days=i) for i in range(days)]
    else:
        day_labels = [f"D{d:02d}" for d in range(1, days + 1)]

    rows: List[Dict[str, Any]] = []

    for di in range(days):
        day = day_labels[di]
        day_row_start = len(rows)
        available = formation_ids.copy()  # 使える編成
        rng = random.Random(seed + di)  # 日ごとに乱数固定
        assigned_inspection_b_ops: set[str] = set()

        # required運用の未割当リスト（その日ごとに消し込む）
        req_today = required_ops.copy()

        def compute_deadhead(
            prev_loc: Any, start_loc: Any, prev_op: Any, op_id: str
        ) -> int:
            # 基本：前日の終点 != 当日の始点 なら回送
            deadhead = int(prev_loc != start_loc)

            # 例外1：JB33終 → 翌日JB30/JB35始 は回送扱いしない
            if deadhead and (prev_loc, start_loc) in NON_DEADHEAD_LOC_JUMPS:
                deadhead = 0

            # 例外2：前日53B → 翌日55B は回送扱いしない（位置が違っていても）
            if deadhead and (str(prev_op), op_id) in NON_DEADHEAD_OP_CHAINS:
                deadhead = 0
            return deadhead

        def can_assign_distB(formation_id: str, op_row) -> bool:
            daysB_before = state[formation_id].daysB
            distB_before = state[formation_id].distB_km
            distance_km = int(op_row.distance_km)
            did_inspection_B = int(
                daysB_before >= 20 and int(op_row.is_inspection_B) == 1
            )
            distB_after = 0 if did_inspection_B else distB_before + distance_km
            return distB_after <= max_distance_since_inspectionB

        def record_assignment(
            formation_id: str,
            op_id: str,
            op_start_loc: Any,
            op_end_loc: Any,
            is_inspection_B: int,
            status: str,
            remove_required: bool,
        ) -> None:
            """1編成に1運用を割当して、rows追加・state更新・availableから除外・req_todayから除外"""
            nonlocal req_today  # 関数外の変数を更新するためnonlocalをつける
            prev_loc = to_station_code(state[formation_id].loc)
            start_loc = to_station_code(op_start_loc)
            prev_op = state[formation_id].prev_op_id

            deadhead = compute_deadhead(prev_loc, start_loc, prev_op, op_id)

            # 検査期限とオーバーの設定
            daysA_before = state[formation_id].daysA
            daysB_before = state[formation_id].daysB
            distB_before = state[formation_id].distB_km
            total_before = state[formation_id].total_km
            did_inspection_A = int(
                daysA_before in (6, 7) and start_loc in inspection_a_start_codes
            )
            did_inspection_B = int(daysB_before >= 20 and int(is_inspection_B) == 1)
            distance_km = int(op_by_id[op_id].distance_km) if op_id in op_by_id else 0

            # 運用後に距離を更新（B検査成立ならリセット）
            if did_inspection_B:
                distB_after = 0
            else:
                distB_after = distB_before + distance_km
            total_after = total_before + distance_km

            rows.append(
                dict(
                    day=day,
                    formation_id=formation_id,
                    operation_id=op_id,
                    status=status,
                    op_start_loc=op_start_loc,
                    op_end_loc=op_end_loc,
                    deadhead=deadhead,
                    distance_km=distance_km,
                    daysA_before=daysA_before,
                    daysB_before=daysB_before,
                    distB_km_before=distB_before,
                    total_km_before=total_before,
                    daysA_after=None,
                    daysB_after=None,
                    distB_km_after=distB_after,
                    total_km_after=total_after,
                    overdueA=None,
                    overdueB=None,
                    did_inspection_A=did_inspection_A,
                    did_inspection_B=did_inspection_B,
                )
            )

            # 状態更新（検査成立→リセット→距離加算済み）
            state[formation_id].loc = op_end_loc
            state[formation_id].prev_op_id = op_id
            if did_inspection_A:
                state[formation_id].daysA = 0
            if did_inspection_B:
                state[formation_id].daysB = 0
            state[formation_id].distB_km = distB_after
            state[formation_id].total_km = total_after
            if formation_id in available:
                available.remove(formation_id)
            if int(is_inspection_B) == 1:
                if op_id in assigned_inspection_b_ops:
                    raise ValueError(f"[{day}] 検査B運用が重複: {op_id}")
                assigned_inspection_b_ops.add(op_id)
            if remove_required:
                # req_todayからop_idを一件だけ消す
                req_today = req_today[req_today["operation_id"] != op_id].reset_index(
                    drop=True
                )

        def assign_one(formation_id: str, op_row) -> None:
            record_assignment(
                formation_id=formation_id,
                op_id=str(op_row.operation_id),
                op_start_loc=op_row.start_loc,
                op_end_loc=op_row.end_loc,
                is_inspection_B=int(op_row.is_inspection_B),
                status="RUN",
                remove_required=True,
            )

        # 0) 翌日固定運用（47B→49B 等）を最優先で割当
        fixed_targets = {}
        req_ids = set(req_today["operation_id"].astype(str))
        for formation_id in list(available):
            prev = state[formation_id].prev_op_id
            if prev in FIXED_NEXT_OP:
                next_op = FIXED_NEXT_OP[prev]
                if next_op in fixed_targets.values():
                    raise ValueError(f"[{day}] 翌日固定が競合: {next_op}")
                if next_op not in req_ids:
                    raise ValueError(
                        f"[{day}] 翌日固定の対象運用がrequiredに存在しない: {prev}→{next_op}"
                    )
                op_row = op_by_id.get(next_op)
                if op_row is None:
                    raise ValueError(
                        f"[{day}] 翌日固定の対象運用がmaster_dataに存在しない: {next_op}"
                    )
                fixed_targets[formation_id] = op_row

        fixed_items = list(fixed_targets.items())
        rng.shuffle(fixed_items)
        for formation_id, op_row in fixed_items:
            assign_one(formation_id, op_row)

        # 1) 検査Bの割当（distBが上限に近い編成を優先）
        b_ops = ops[ops["is_inspection_B"] == 1].copy()
        if not b_ops.empty:
            b_ops = b_ops[
                ~b_ops["operation_id"].astype(str).isin(assigned_inspection_b_ops)
            ]

        eligible_formations = [
            formation_id
            for formation_id in available
            if state[formation_id].daysB >= 20
        ]  # B検査から20日以上経過した編成をリストアップ

        if not b_ops.empty and eligible_formations:
            model = cp_model.CpModel()
            b_op_ids = list(b_ops["operation_id"].astype(str))
            y: Dict[Tuple[str, str], cp_model.IntVar] = {}
            for formation_id in eligible_formations:
                for op_id in b_op_ids:
                    y[(formation_id, op_id)] = model.NewBoolVar(
                        f"y_{formation_id}_{op_id}"
                    )  # formation_id を、B検査運用 op_id に割り当てるなら 1、割り当てないなら 0

            for formation_id in eligible_formations:
                model.Add(
                    sum(y[(formation_id, op_id)] for op_id in b_op_ids) <= 1
                )  # 同じ編成が複数のB運用に同時に入るのはダメ

            for op_id in b_op_ids:
                model.Add(
                    sum(
                        y[(formation_id, op_id)] for formation_id in eligible_formations
                    )
                    <= 1
                )  # 同じB運用枠に複数編成を突っ込むのはダメ

            # distB/日数が大きい編成ほど優先（期限・距離の逼迫を優先）
            # 併せて start_loc が一致する割当を優先し、deadhead を減らす。
            model.Maximize(
                sum(
                    y[(formation_id, op_id)]
                    * (
                        state[formation_id].distB_km * 10
                        + state[formation_id].daysB * 5
                        + (
                            1
                            if state[formation_id].loc == op_by_id[op_id].start_loc
                            else 0
                        )
                    )
                    for formation_id in eligible_formations
                    for op_id in b_op_ids
                )
            )

            solver = cp_model.CpSolver()
            status = solver.Solve(model)
            if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
                assignments = []
                for formation_id in eligible_formations:
                    for op_id in b_op_ids:
                        if solver.Value(y[(formation_id, op_id)]) == 1:
                            assignments.append((formation_id, op_id))
                            # 確定した組み合わせに割り当て
                for formation_id, op_id in assignments:
                    op_row = op_by_id.get(op_id)
                    if op_row is None:
                        raise ValueError(
                            f"[{day}] 検査Bの対象運用がmaster_dataに存在しない: {op_id}"
                        )
                    assign_one(formation_id, op_row)

        # 3) 残り required を CP-SAT で割当（距離偏り最小＋deadhead抑制＋期限超過抑制）
        if len(req_today) > 0 and available:
            model = cp_model.CpModel()
            req_list = list(req_today.itertuples(index=False))
            req_ids = [str(op.operation_id) for op in req_list]
            max_distance_today = (
                max(int(op.distance_km) for op in req_list) if req_list else 0
            )

            y: Dict[Tuple[str, str], cp_model.IntVar] = {}
            deadhead_cost: Dict[Tuple[str, str], int] = {}
            overdue_cost: Dict[Tuple[str, str], int] = {}
            distance_map: Dict[Tuple[str, str], int] = {}

            for formation_id in available:
                f_state = state[formation_id]
                prev_loc = to_station_code(f_state.loc)
                prev_op = f_state.prev_op_id
                daysA_before = f_state.daysA
                daysB_before = f_state.daysB
                distB_before = f_state.distB_km

                overdueA_idle = int(daysA_before + 1 >= max_days_since_inspectionA + 1)
                overdueB_idle = int(daysB_before + 1 >= max_days_since_inspectionB + 1)

                for op in req_list:
                    op_id = str(op.operation_id)
                    start_loc = to_station_code(op.start_loc)
                    distance_km = int(op.distance_km)
                    is_inspection_B = int(op.is_inspection_B)

                    did_inspection_B = int(daysB_before >= 20 and is_inspection_B == 1)
                    if did_inspection_B:
                        distB_after = 0
                    else:
                        distB_after = distB_before + distance_km
                    if distB_after > max_distance_since_inspectionB:
                        continue

                    y[(formation_id, op_id)] = model.NewBoolVar(
                        f"y_{formation_id}_{op_id}"
                    )
                    distance_map[(formation_id, op_id)] = distance_km

                    did_inspection_A = int(
                        daysA_before in (6, 7) and start_loc in inspection_a_start_codes
                    )
                    daysA_after = 0 if did_inspection_A else daysA_before + 1
                    daysB_after = 0 if did_inspection_B else daysB_before + 1
                    overdueA_after = int(daysA_after >= max_days_since_inspectionA + 1)
                    overdueB_after = int(daysB_after >= max_days_since_inspectionB + 1)
                    overdue_cost[(formation_id, op_id)] = (
                        overdueA_idle
                        + overdueB_idle
                        + (overdueA_after - overdueA_idle)
                        + (overdueB_after - overdueB_idle)
                    )
                    deadhead_cost[(formation_id, op_id)] = compute_deadhead(
                        prev_loc, start_loc, prev_op, op_id
                    )

            for formation_id in available:
                model.Add(
                    sum(
                        y[(formation_id, op_id)]
                        for op_id in req_ids
                        if (formation_id, op_id) in y
                    )
                    <= 1
                )

            for op_id in req_ids:
                candidates = [
                    y[(formation_id, op_id)]
                    for formation_id in available
                    if (formation_id, op_id) in y
                ]
                if not candidates:
                    raise ValueError(
                        f"[{day}] required運用をdistB制約内で割当できません: {op_id}"
                    )
                model.Add(sum(candidates) == 1)

            total_after_vars: Dict[str, cp_model.IntVar] = {}
            for formation_id in available:
                f_state = state[formation_id]
                total_before = f_state.total_km
                total_after = model.NewIntVar(
                    total_before,
                    total_before + max_distance_today,
                    f"total_{formation_id}",
                )
                model.Add(
                    total_after
                    == total_before
                    + sum(
                        y[(formation_id, op_id)] * distance_map[(formation_id, op_id)]
                        for op_id in req_ids
                        if (formation_id, op_id) in y
                    )
                )
                total_after_vars[formation_id] = total_after

            total_all: List[cp_model.IntVar] = []
            for formation_id in formation_ids:
                if formation_id in total_after_vars:
                    total_all.append(total_after_vars[formation_id])
                else:
                    total_const = model.NewIntVar(
                        state[formation_id].total_km,
                        state[formation_id].total_km,
                        f"total_{formation_id}_const",
                    )
                    total_all.append(total_const)

            total_min_bound = min(state[fid].total_km for fid in formation_ids)
            total_max_bound = max(state[fid].total_km for fid in formation_ids) + (
                max_distance_today if req_list else 0
            )
            max_total = model.NewIntVar(total_min_bound, total_max_bound, "max_total")
            min_total = model.NewIntVar(total_min_bound, total_max_bound, "min_total")
            model.AddMaxEquality(max_total, total_all)
            model.AddMinEquality(min_total, total_all)

            model.Minimize(
                W_RANGE * (max_total - min_total)
                + W_DEADHEAD
                * sum(
                    y[(formation_id, op_id)] * deadhead_cost[(formation_id, op_id)]
                    for formation_id in available
                    for op_id in req_ids
                    if (formation_id, op_id) in y
                )
                + W_OVD
                * sum(
                    y[(formation_id, op_id)] * overdue_cost[(formation_id, op_id)]
                    for formation_id in available
                    for op_id in req_ids
                    if (formation_id, op_id) in y
                )
            )

            solver = cp_model.CpSolver()
            status = solver.Solve(model)
            if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
                raise ValueError(f"[{day}] required運用のCP-SAT割当が不可能です")

            for formation_id in available.copy():
                for op_id in req_ids:
                    if (formation_id, op_id) in y and solver.Value(
                        y[(formation_id, op_id)]
                    ) == 1:
                        op_row = op_by_id.get(op_id)
                        if op_row is None:
                            raise ValueError(
                                f"[{day}] required運用がmaster_dataに存在しない: {op_id}"
                            )
                        assign_one(formation_id, op_row)

        # 余り編成は待機運用へ
        for formation_id in list(available):
            loc = state[formation_id].loc
            idle_op = idle_by_start.get(loc, default_idle_op)
            idle_row = op_by_id.get(idle_op, None)

            record_assignment(
                formation_id=formation_id,
                op_id=str(idle_op),
                op_start_loc=(idle_row.start_loc if idle_row else loc),
                op_end_loc=(idle_row.end_loc if idle_row else loc),
                is_inspection_B=int(idle_row.is_inspection_B) if idle_row else 0,
                status="IDLE",
                remove_required=False,
            )

        # 日末更新：検査成立済み状態に対して日数+1し、運用終了後の値を記録
        for formation_id in formation_ids:
            state[formation_id].daysA += 1
            state[formation_id].daysB += 1

        for row in rows[day_row_start:]:
            fid = row["formation_id"]
            row["daysA_after"] = state[fid].daysA
            row["daysB_after"] = state[fid].daysB
            row["distB_km_after"] = state[fid].distB_km
            row["total_km_after"] = state[fid].total_km
            row["overdueA"] = int(state[fid].daysA >= max_days_since_inspectionA + 1)
            row["overdueB"] = int(state[fid].daysB >= max_days_since_inspectionB + 1)

    schedule = (
        pd.DataFrame(rows).sort_values(["day", "formation_id"]).reset_index(drop=True)
    )

    schedule = schedule.rename(
        columns={
            "daysA_after": "daysA",
            "daysB_after": "daysB",
            "distB_km_after": "distB_km",
            "total_km_after": "total_km",
        }
    )
    drop_cols = [
        "daysA_before",
        "daysB_before",
        "distB_km_before",
        "total_km_before",
        "daysA_after",
        "daysB_after",
        "distB_km_after",
        "total_km_after",
    ]
    schedule = schedule.drop(columns=[c for c in drop_cols if c in schedule.columns])

    # ガント風：編成×日 の行列
    gantt_ops = schedule.pivot(
        index="formation_id", columns="day", values="operation_id"
    )

    # 位置ガント：start/end を別pivotで持つ（Excel側で1日2列にする）
    gantt_start = schedule.pivot(
        index="formation_id", columns="day", values="op_start_loc"
    )
    gantt_end = schedule.pivot(index="formation_id", columns="day", values="op_end_loc")

    return (
        schedule,
        gantt_ops.reset_index(),
        gantt_start,
        gantt_end,
    )


# ---------- Excel export ----------


def add_sheet_from_df(
    wb: openpyxl.Workbook,
    name: str,
    df: pd.DataFrame,
    table_name: Optional[str] = None,
    freeze: str = "A2",
    station_cols: Optional[List[str]] = None,
) -> None:
    ws = wb.create_sheet(name)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # header style
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # col widths (rough)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for c in col[:200]:
            v = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 40)

    ws.freeze_panes = freeze
    # フィルターだけ付ける
    ws.auto_filter.ref = ws.dimensions

    # 駅カラムに色付け（指定された列だけ）
    if station_cols:
        header = [c.value for c in ws[1]]
        col_idx = {str(v): i + 1 for i, v in enumerate(header) if v is not None}
        for col_name in station_cols:
            if col_name not in col_idx:
                continue
            j = col_idx[col_name]
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, j)
                fill = station_fill(cell.value)
                if fill:
                    cell.value = to_station_code(cell.value)
                    cell.fill = fill
                    set_white_font(cell, cell.value)


def add_gantt_loc_sheet(
    wb: openpyxl.Workbook,
    name: str,
    gantt_start: pd.DataFrame,
    gantt_end: pd.DataFrame,
) -> None:
    ws = wb.create_sheet(name)
    days = list(gantt_start.columns)

    # Row1は2列分書き込む。D01→None→D02→Noneの順
    ws.cell(1, 1, "formation_id")
    c = 2
    for d in days:
        ws.cell(1, c, d)
        ws.cell(1, c + 1, None)
        c += 2

    # Row2はstartとendを交互に書く
    ws.cell(2, 1, None)
    c = 2
    for _ in days:
        ws.cell(2, c, "start")
        ws.cell(2, c + 1, "end")
        c += 2

    # Header style（2段とも）
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for r in (1, 2):
        for cell in ws[r]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    start_df = gantt_start.copy()
    end_df = gantt_end.copy()
    start_df.index = start_df.index.astype(str)
    end_df.index = end_df.index.astype(str)

    out_r = 3  # 3行目から
    for formation_id in start_df.index:
        ws.cell(out_r, 1, formation_id)
        c = 2
        for d in days:
            sv = to_station_code(start_df.loc[formation_id, d])
            ev = to_station_code(end_df.loc[formation_id, d])
            ws.cell(out_r, c, sv)
            ws.cell(out_r, c + 1, ev)

            # 色付け（start/endとも）
            for cc, v in [(c, sv), (c + 1, ev)]:
                fill = station_fill(v)
                if fill:
                    cell = ws.cell(out_r, cc)
                    cell.fill = fill
                    set_white_font(cell, v)
            c += 2
        out_r += 1

    # 幅調整とfreeze
    ws.freeze_panes = "B3"  # ウィンドウ枠の固定
    ws.column_dimensions["A"].width = 14
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 8


def add_formation_triplet_sheet(
    wb: openpyxl.Workbook,
    name: str,
    schedule: pd.DataFrame,
    days: list[str],
    formations: list[str],
) -> None:
    ws = wb.create_sheet(name)

    # Row1: 編成ヘッダ（3列結合）
    ws.cell(1, 1, None)
    col = 2
    for formation_id in formations:
        ws.cell(1, col, formation_id)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        col += 3

    # Row2: start/end/operation
    ws.cell(2, 1, None)
    col = 2
    for _ in formations:
        ws.cell(2, col, "start")
        ws.cell(2, col + 1, "end")
        ws.cell(2, col + 2, "operation")
        col += 3

    # ヘッダスタイル（2段）
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for r in (1, 2):
        for cell in ws[r]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 参照しやすいように pivot
    schedule_copy = schedule.copy()
    schedule_copy["day"] = schedule_copy["day"].astype(str)
    schedule_copy["formation_id"] = schedule_copy["formation_id"].astype(str)

    p_start = schedule_copy.pivot(
        index="day", columns="formation_id", values="op_start_loc"
    ).reindex(index=days, columns=formations)
    p_end = schedule_copy.pivot(
        index="day", columns="formation_id", values="op_end_loc"
    ).reindex(index=days, columns=formations)
    p_op = schedule_copy.pivot(
        index="day", columns="formation_id", values="operation_id"
    ).reindex(index=days, columns=formations)
    p_dead = schedule_copy.pivot(
        index="day", columns="formation_id", values="deadhead"
    ).reindex(index=days, columns=formations)

    # Data rows
    out_r = 3  # 3行目から
    for day in days:
        ws.cell(out_r, 1, day)
        ws.cell(out_r, 1).alignment = Alignment(horizontal="center", vertical="center")

        col = 2
        for formation_id in formations:
            sv = to_station_code(p_start.loc[day, formation_id])
            ev = to_station_code(p_end.loc[day, formation_id])
            ov = p_op.loc[day, formation_id]
            dv = p_dead.loc[day, formation_id]

            c_start = ws.cell(out_r, col, sv)
            c_end = ws.cell(out_r, col + 1, ev)
            c_op = ws.cell(out_r, col + 2, (None if pd.isna(ov) else str(ov)))

            # start/end は駅色を塗る（既存ルールに合わせる）
            for cell, v in [(c_start, sv), (c_end, ev)]:
                fill = station_fill(v)
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # operation は中央寄せ
            c_op.alignment = Alignment(horizontal="center", vertical="center")

            # deadhead=1 のとき：startセルを「赤字＋太字」
            if pd.notna(dv) and int(dv) == 1:
                # 既存フォント設定は引き継いで、色と太字だけ上書き
                c_start.font = c_start.font.copy(color="FF0000", bold=True)

            col += 3
        out_r += 1

    # --- 3列セット（start/end/op）の境界を太線にする（各ブロック最終列の右側） ---
    THICK = Side(style="thick", color="000000")
    row_max = ws.max_row
    start_col = 2
    cols_per = 3
    for i in range(len(formations)):
        boundary_col = start_col + cols_per * (i + 1) - 1
        for r in range(1, row_max + 1):
            cell = ws.cell(r, boundary_col)
            b = cell.border
            cell.border = Border(
                left=b.left,
                right=THICK,
                top=b.top,
                bottom=b.bottom,
                diagonal=b.diagonal,
                diagonal_direction=b.diagonal_direction,
                outline=b.outline,
                vertical=b.vertical,
                horizontal=b.horizontal,
            )

    # --- 行方向（横方向）に細い罫線：各セルの bottom を thin にする ---
    THIN = Side(style="thin", color="000000")
    col_max = ws.max_column
    for r in range(1, row_max + 1):
        for c in range(1, col_max + 1):
            cell = ws.cell(r, c)
            b = cell.border
            cell.border = Border(
                left=b.left,
                right=b.right,
                top=b.top,
                bottom=THIN,
                diagonal=b.diagonal,
                diagonal_direction=b.diagonal_direction,
                outline=b.outline,
                vertical=b.vertical,
                horizontal=b.horizontal,
            )

    # 先頭行の上端も線が欲しい場合（任意）
    for c in range(1, col_max + 1):
        cell = ws.cell(1, c)
        b = cell.border
        cell.border = Border(
            left=b.left,
            right=b.right,
            top=THIN,
            bottom=b.bottom,
            diagonal=b.diagonal,
            diagonal_direction=b.diagonal_direction,
            outline=b.outline,
            vertical=b.vertical,
            horizontal=b.horizontal,
        )

    # 罫線っぽく見せる（列幅と固定）
    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 8
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10


def add_inspection_triplet_sheet(
    wb: openpyxl.Workbook,
    name: str,
    schedule: pd.DataFrame,
    days: list[str],
    formations: list[str],
) -> None:
    ws = wb.create_sheet(name)

    ws.cell(1, 1, None)
    col = 2
    for formation_id in formations:
        ws.cell(1, col, formation_id)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
        col += 4

    ws.cell(2, 1, None)
    col = 2
    for _ in formations:
        ws.cell(2, col, "operation")
        ws.cell(2, col + 1, "daysA")
        ws.cell(2, col + 2, "daysB")
        ws.cell(2, col + 3, "distB_km")
        col += 4

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for r in (1, 2):
        for cell in ws[r]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    schedule_copy = schedule.copy()
    schedule_copy["day"] = schedule_copy["day"].astype(str)
    schedule_copy["formation_id"] = schedule_copy["formation_id"].astype(str)

    p_op = schedule_copy.pivot(
        index="day", columns="formation_id", values="operation_id"
    ).reindex(index=days, columns=formations)
    p_daysA = schedule_copy.pivot(
        index="day", columns="formation_id", values="daysA"
    ).reindex(index=days, columns=formations)

    p_daysB = schedule_copy.pivot(
        index="day", columns="formation_id", values="daysB"
    ).reindex(index=days, columns=formations)

    p_distB = schedule_copy.pivot(
        index="day", columns="formation_id", values="distB_km"
    ).reindex(index=days, columns=formations)

    out_r = 3
    for day in days:
        ws.cell(out_r, 1, day)
        ws.cell(out_r, 1).alignment = Alignment(horizontal="center", vertical="center")

        col = 2
        for formation_id in formations:
            ov = p_op.loc[day, formation_id]
            av = p_daysA.loc[day, formation_id]
            bv = p_daysB.loc[day, formation_id]
            dv = p_distB.loc[day, formation_id]

            c_op = ws.cell(out_r, col, (None if pd.isna(ov) else str(ov)))
            c_a = ws.cell(out_r, col + 1, (None if pd.isna(av) else int(av)))
            c_b = ws.cell(out_r, col + 2, (None if pd.isna(bv) else int(bv)))
            c_d = ws.cell(out_r, col + 3, (None if pd.isna(dv) else int(dv)))

            for cell in (c_op, c_a, c_b, c_d):
                cell.alignment = Alignment(horizontal="center", vertical="center")

            col += 4
        out_r += 1

    THICK = Side(style="thick", color="000000")
    row_max = ws.max_row
    start_col = 2
    cols_per = 4
    for i in range(len(formations)):
        boundary_col = start_col + cols_per * (i + 1) - 1
        for r in range(1, row_max + 1):
            cell = ws.cell(r, boundary_col)
            b = cell.border
            cell.border = Border(
                left=b.left,
                right=THICK,
                top=b.top,
                bottom=b.bottom,
                diagonal=b.diagonal,
                diagonal_direction=b.diagonal_direction,
                outline=b.outline,
                vertical=b.vertical,
                horizontal=b.horizontal,
            )

    THIN = Side(style="thin", color="000000")
    col_max = ws.max_column
    for r in range(1, row_max + 1):
        for c in range(1, col_max + 1):
            cell = ws.cell(r, c)
            b = cell.border
            cell.border = Border(
                left=b.left,
                right=b.right,
                top=b.top,
                bottom=THIN,
                diagonal=b.diagonal,
                diagonal_direction=b.diagonal_direction,
                outline=b.outline,
                vertical=b.vertical,
                horizontal=b.horizontal,
            )

    for c in range(1, col_max + 1):
        cell = ws.cell(1, c)
        b = cell.border
        cell.border = Border(
            left=b.left,
            right=b.right,
            top=THIN,
            bottom=b.bottom,
            diagonal=b.diagonal,
            diagonal_direction=b.diagonal_direction,
            outline=b.outline,
            vertical=b.vertical,
            horizontal=b.horizontal,
        )

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 8
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10


# Excel出力
def export_baseline_excel(
    out_path: str,
    schedule: pd.DataFrame,
    gantt_ops: pd.DataFrame,
    gantt_start: pd.DataFrame,
    gantt_end: pd.DataFrame,
    ops: pd.DataFrame,
    forms: pd.DataFrame,
) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    add_sheet_from_df(
        wb,
        "schedule_long",
        schedule,
        table_name="ScheduleLong",
        freeze="A2",
        station_cols=["op_start_loc", "op_end_loc"],
    )
    add_sheet_from_df(wb, "gantt_ops", gantt_ops, table_name="GanttOps", freeze="B2")
    add_gantt_loc_sheet(wb, "gantt_loc", gantt_start, gantt_end)
    # day/formation の順序をここで固定
    days = sorted(schedule["day"].astype(str).unique(), key=lambda s: (len(s), s))
    formations = sorted(schedule["formation_id"].astype(str).unique())
    add_formation_triplet_sheet(wb, "gantt_triplet", schedule, days, formations)
    add_inspection_triplet_sheet(
        wb, "gantt_inspection_triplet", schedule, days, formations
    )
    add_sheet_from_df(
        wb,
        "master_operations",
        ops,
        table_name="MasterOps",
        freeze="A2",
        station_cols=["start_loc", "end_loc"],
    )
    add_sheet_from_df(
        wb,
        "master_formations",
        forms,
        table_name="MasterForms",
        freeze="A2",
        station_cols=["init_location"],
    )

    wb.save(out_path)


# ---------- CLI ----------


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--master", default="data/master_data.xlsx", help="master_data.xlsx のパス"
    )
    ap.add_argument(
        "--out", default="outputs/baseline_output.xlsx", help="出力Excelパス"
    )
    ap.add_argument("--days", type=int, default=30, help="何日分作るか")
    ap.add_argument(
        "--start-date",
        default=None,
        help="YYYY-MM-DD を指定すると日付ラベルになる（未指定なら D01..）",
    )
    ap.add_argument(
        "--default-idle-op",
        default="IDOL_Mitaka",
        help="待機運用が引けない時のデフォルト operation_id",
    )
    ap.add_argument(
        "--seed", type=int, default=42, help="津田沼ランダム割当の乱数seed（再現性用）"
    )
    ap.add_argument(
        "--tsudanuma-code",
        default="JB33",
        help="津田沼の駅コード（master_dataに合わせる）",
    )

    args = ap.parse_args()

    ops, forms, rules = read_master(args.master)
    schedule, gantt_ops, gantt_start, gantt_end = make_baseline_schedule(
        ops,
        forms,
        max_days_since_inspectionA=rules["max_days_since_inspectionA"],
        max_days_since_inspectionB=rules["max_days_since_inspectionB"],
        max_distance_since_inspectionB=rules["max_distance_since_inspectionB"],
        days=args.days,
        start_date=args.start_date,
        default_idle_op=args.default_idle_op,
        seed=args.seed,
        tsudanuma_code=args.tsudanuma_code,
    )
    export_baseline_excel(
        args.out, schedule, gantt_ops, gantt_start, gantt_end, ops, forms
    )
    print(f"OK: {args.out}")
    print(f"schedule_long: {schedule.shape[0]} rows, {schedule.shape[1]} cols")


if __name__ == "__main__":
    main()
